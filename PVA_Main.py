import tkinter
import tkinter as tk
from datetime import date
from tkinter import *
from tkinter import filedialog
import os
import openpyxl
from openpyxl import *
import warnings
import ctypes

global Est
global hubble 
global JDE_check
Est = False
hubble = False
JDE_check = False

def run_analysis():
    #get WO number from the text feild
    WO_number = WO_number_Entry.get()

    #check if WO number entered is valid if not display an error box
    while True:
        if WO_number.isdigit() or len(str(WO_number)) == 6:
            WO_number = int(WO_number)
            break 
        else: 
            ctypes.windll.user32.MessageBoxW(0, "Enter a 6 digit WO number", "Invalid input", )
            return None
        

    # Load the PVA template to write on
    PVA_WB = openpyxl.load_workbook('PVA_Template.xlsx')
    PVA_WS = PVA_WB.active

    #load P6 extract workbook
    P6_extract = openpyxl.load_workbook('P6_Extract.xlsx')
    P6_WS = P6_extract.active

    #update WO info from P6 extract
    for i in range (1, P6_WS.max_row + 1):
        WO = P6_WS.cell(row= i, column= 1).value
        if WO == WO_number:
            PVA_WS['B5'] = P6_WS.cell(row= i, column= 2).value
            PVA_WS['B6'] = P6_WS.cell(row= i, column= 12).value


    # update date and WO number on template
    PVA_WS['B4'].value = WO_number
    PVA_WS['H2'].value = date.today()

    global Estimates_WB
    global Estimates_file
    global Actuals_WB_Hubble
    global Actuals_WB_JDE

    
    tasks = ['130', '131', '133', '200', '205', '225', '235', '245', '255', '270', '280', '295', '315', '380', '400', '415', '430', '440', '450', '455', '465', '485', '500', '510', '515', '599', '999']
    P_C = ['600', '610', '620', '630', '640', '650', '660', '670', '680', '690', '700', '710', '715', '720', '725', '730']


    if Estimates_WB is not None:
        Estimates_WS = Estimates_WB.active
        
        mat = 0
        mat_brdn = 0 
        mat_dir = 0
        personnel = 0
        personnel_brdn = 0
        travel = 0
        tools = 0
        ex_cont = 0
        ex_cont_brdn = 0
        other = 0
        ICI_Credit = 0
        additional_est = 0
        ins_hours = 0
        ins_hours_OT = 0
        lines_hours = 0
        lines_hours_OT = 0
        PC_hours = 0
        PC_hours_OT = 0
        des_hours = 0
        des_hours_OT = 0
        met_hours = 0
        met_hours_OT = 0

    
        for i in range (18, 36):
            if i == 21:
                continue
            if i == 25:
                continue
            if i == 27:
                continue
            for k in range (2, 4):
                if k == 3:
                    continue
                PVA_WS.cell(row= i, column= k ).value = None

        for i in range (44, 55):
            for k in range (2, 3):
                PVA_WS.cell(row= i, column= k ).value = None

        for i in range (1, Estimates_WS.max_row + 1):

            sub = Estimates_WS.cell(row= i, column= 12).value
            des = Estimates_WS.cell(row= i, column= 13).value
            Wdes = Estimates_WS.cell(row= i, column= 17).value
            dollars = Estimates_WS.cell(row= i, column= 32).value
            brdn = Estimates_WS.cell(row= i, column= 34).value
            fringe = Estimates_WS.cell(row= i, column= 35).value
            hours = Estimates_WS.cell(row= i, column= 28).value
            hours_OT = Estimates_WS.cell(row= i, column= 29).value


            if sub == 1  and des == 'Direct Materials':
                mat_dir = dollars
               
            elif sub == 1 and des == "Miscellaneous Cost":
                ex_cont = dollars
                if brdn != None or fringe != None:
                    ex_cont_brdn = brdn

            elif sub == 1:
                mat_brdn = brdn
                mat = dollars
                if fringe != None:
                    mat_brdn += fringe
                
            if sub == 100 and Wdes == 'Travel Time':
                travel = dollars 
                if brdn != None:
                    travel += brdn
                if fringe != None:
                    travel += fringe
    
            elif sub == 100:
                tools += dollars
                if hours_OT != None:
                        lines_hours_OT += hours_OT
            if sub == 199:
                ICI_Credit = dollars

            for k in range(len(tasks)):
                if sub == float(tasks[k]):
                    personnel += dollars
                    lines_hours += hours
                    if brdn != None:
                        personnel_brdn += brdn
                    if fringe != None:
                        personnel_brdn += fringe
                    if hours_OT != None:
                        lines_hours_OT += hours_OT

            for k in range(len(P_C)):
                if sub == float(P_C[k]):
                    personnel += dollars
                    PC_hours += hours
                    if brdn != None:
                        personnel_brdn += brdn
                    if fringe != None:
                        personnel_brdn += fringe
                    if hours_OT != None:
                        PC_hours_OT += hours_OT

            if sub == 170 or sub == 115:
                des_hours += hours
                personnel += dollars
                if brdn != None:
                    personnel_brdn += brdn
                if fringe != None:
                    personnel_brdn += fringe
                if hours_OT != None:
                    des_hours_OT += hours_OT
                        
            if sub == 120:
                ins_hours = hours
                personnel += dollars
                if hours_OT != None:
                    ins_hours_OT = hours_OT
                if brdn != None:
                    personnel_brdn += brdn
                if fringe != None:
                    personnel_brdn += fringe
                        
            if sub == 800:
                met_hours = hours
                if hours_OT != None:
                    met_hours_OT = hours_OT

            PVA_WS['B18'] = mat
            PVA_WS['B19'] = mat_brdn
            PVA_WS['B20'] = mat_dir
            PVA_WS['B22'] = personnel
            PVA_WS['B23'] = personnel_brdn 
            PVA_WS['B24'] = travel
            PVA_WS['B26'] = tools
            PVA_WS['B28'] = ex_cont
            PVA_WS['B29'] = ex_cont_brdn
            PVA_WS['B30'] = other
            PVA_WS['B31'] = ICI_Credit
            PVA_WS['B32'] = additional_est
            PVA_WS['B46'] = des_hours
            PVA_WS['B47'] = des_hours_OT
            PVA_WS['B48'] = ins_hours
            PVA_WS['B49'] = ins_hours_OT
            PVA_WS['B50'] = lines_hours
            PVA_WS['B51'] = lines_hours_OT
            PVA_WS['B52'] = PC_hours
            PVA_WS['B53'] = PC_hours_OT
            PVA_WS['B54'] = met_hours
            PVA_WS['B55'] = met_hours_OT
            

    if Actuals_WB_JDE is not None:
        Hours_actual_type1 = 0
        Hours_actual_type1_OT = 0
        Hours_actual_des = 0
        Hours_actual_des_OT = 0
        Hours_actual_ins = 0
        Hours_actual_ins_OT = 0
        Hours_actual_lines = 0
        Hours_actual_lines_OT = 0
        Hours_actual_PC = 0
        Hours_actual_PC_OT = 0
        Hours_actual_met = 0
        Hours_actual_met_OT = 0
        

        Actuals_WS_JDE = Actuals_WB_JDE.active
        
        for i in range(1, Actuals_WS_JDE.max_row + 1):
            if Actuals_WS_JDE.cell(row= i, column= 1).value == "Labor Billing Distribution":
                start = i
            if Actuals_WS_JDE.cell(row= i, column= 1).value == "Total:Labor Billing Distribution":
                end = i
                break

        for i in range(start, end):
            code = Actuals_WS_JDE.cell(row= i, column= 7).value
            hours_actual = Actuals_WS_JDE.cell(row= i, column= 4).value
            hour_desc = Actuals_WS_JDE.cell(row= i, column= 1).value

            if code != None:
                code_split = code.split(".")
                task_code = code_split[-1]
            else:
                task_code = 0
            if hour_desc != None:
                desc_split = hour_desc.split()
            if len(desc_split) >= 4:
                hour_type = desc_split[-2]
            if hours_actual != None:
                if task_code == '1':
                    if hour_type == "003" or hour_type == "322" or hour_type == "086" or hour_type == "045":
                        Hours_actual_type1_OT += hours_actual
                    else:
                        Hours_actual_type1 += hours_actual
                if task_code == '170' or task_code == '115' or task_code == '110':
                    if hour_type == "003" or hour_type == "322" or hour_type == "086" or hour_type == "045":
                        Hours_actual_des_OT += hours_actual
                    else:
                        Hours_actual_des += hours_actual
                if task_code == '120':
                    if hour_type == "003" or hour_type == "322" or hour_type == "086" or hour_type == "045":
                        Hours_actual_ins_OT += hours_actual
                    else:
                        Hours_actual_ins += hours_actual                        
                if task_code == '100':
                    if hour_type == "003" or hour_type == "322" or hour_type == "086" or hour_type == "045":
                        Hours_actual_lines_OT += hours_actual
                    else:
                        Hours_actual_lines += hours_actual                        
                for k in range(len(tasks)) :
                    if task_code == tasks[k]:
                        if hour_type == "003" or hour_type == "322" or hour_type == "086" or hour_type == "045":
                            Hours_actual_lines_OT += hours_actual
                            break
                        else:
                            Hours_actual_lines += hours_actual
                            break
                for k in range(len(P_C)):
                    if task_code == P_C[k]:
                        if hour_type == "003" or hour_type == "322" or hour_type == "086" or hour_type == "045":
                            Hours_actual_PC_OT += hours_actual
                            break
                        else:
                            Hours_actual_PC += hours_actual
                            break
                if task_code == '800':
                    if hour_type == "003" or hour_type == "322" or hour_type == "086" or hour_type == "045":
                        Hours_actual_met_OT += hours_actual
                    else:
                        Hours_actual_met += hours_actual                        

        PVA_WS['C44'] = Hours_actual_type1
        PVA_WS['C45'] = Hours_actual_type1_OT
        PVA_WS['C46'] = Hours_actual_des
        PVA_WS['C47'] = Hours_actual_des_OT
        PVA_WS['C48'] = Hours_actual_ins
        PVA_WS['C49'] = Hours_actual_ins_OT
        PVA_WS['C50'] = Hours_actual_lines
        PVA_WS['C51'] = Hours_actual_lines_OT
        PVA_WS['C52'] = Hours_actual_PC
        PVA_WS['C53'] = Hours_actual_PC_OT
        PVA_WS['C54'] = Hours_actual_met
        PVA_WS['C55'] = Hours_actual_met_OT

    if Actuals_WB_Hubble is not None:
        Actuals_WS_Hubble = Actuals_WB_Hubble["Template - Summary wo analysis "]
        Brdn_WS_Hubble = Actuals_WB_Hubble["Period Summary"]
        personnel_actuals = 0
        personnel_actuals_brdn = 0
        mat_actuals = 0
        mat_actuals_brdn = 0
        mat_actuals_dirm = 0
        travel_actuals = 0
        tools_actuals = 0
        sc_actuals = 0
        sc_actuals_brdn = 0
        other_actuals = 0
        ici_actuals = 0
        additional_actuals = 0

        for i in range(1, Actuals_WS_Hubble.max_row + 1):
            category = Actuals_WS_Hubble.cell(row= i, column= 1).value
            actual_dollars = Actuals_WS_Hubble.cell(row= i, column= 3).value
            if category != None:
                category_split = category.split(" - ")
                category_type = category_split[-1]
            else:
                category_type = 0
            if category_type == "DESR" or category_type == "LABR" or category_type == "LABO" or category_type == "NUR" or category_type == "NUO":
                personnel_actuals += actual_dollars
            if category_type == "DESF" or category_type == "LABF" or category_type == "LABB":
                personnel_actuals_brdn += actual_dollars
            if category_type == "MATC":
                mat_actuals += actual_dollars
            if category_type == "MATB":
                mat_actuals_brdn += actual_dollars
            if category_type == "DIRM":
                mat_actuals_dirm += actual_dollars
            if category_type == "Fleet":
                tools_actuals += actual_dollars
            if category_type == "Contract Labour":
                sc_actuals += actual_dollars
            if category_type == "Other":
                other_actuals += actual_dollars
            if category_type == "NUR" or category_type == "NUO":
                additional_actuals += actual_dollars
        for i in range(1, Brdn_WS_Hubble.max_row + 1):
            sub_code = Brdn_WS_Hubble.cell(row= i, column= 5).value
            brdn_total = Brdn_WS_Hubble.cell(row= i, column= 8).value
            
            if sub_code == "1":
                sc_actuals_brdn += brdn_total
                personnel_actuals_brdn -= sc_actuals_brdn
            

        PVA_WS['D18'] = mat_actuals
        PVA_WS['D19'] = mat_actuals_brdn
        PVA_WS['D20'] = mat_actuals_dirm
        PVA_WS['D22'] = personnel_actuals
        PVA_WS['D23'] = personnel_actuals_brdn
        PVA_WS['D24'] = travel_actuals
        PVA_WS['D26'] = tools_actuals
        PVA_WS['D28'] = sc_actuals
        PVA_WS['D29'] = sc_actuals_brdn
        PVA_WS['D30'] = other_actuals
        PVA_WS['D31'] = ici_actuals
        PVA_WS['D32'] = additional_actuals


    download_Directory = filedialog.askdirectory()
    file_name = str(WO_number) + " - PVA.xlsx"
    file_path = os.path.join(download_Directory, file_name)
    PVA_WB.save(file_path)
    PVA_WB.close()
    Estimates_WB.close()
    Actuals_WB_Hubble.close()
    Actuals_WB_JDE.close()
    run_button.config(state= 'disabled')

    label_file_explorer_estimates.configure(text="", fg = "blue")
    label_file_explorer_actuals_hubble.configure(text="", fg = "blue")
    label_file_explorer_actuals_JDE.configure(text="", fg = "blue")
    WO_number_Entry.delete(0, END)

def browseFiles_Estimates():
    global Estimates_WB
    global Estimates_file
    global Est
    global hubble 
    global JDE_check
    Est = False


    Estimates_file = filedialog.askopenfilename(initialdir = "/", title = "Select a File",filetypes = (("Excel Workbooks", "*.xlsx*"),("all files", "*.*")))

    # Change label contents
    label_file_explorer_estimates.configure(text="File Opened: "+ Estimates_file, fg = "blue")

    if Estimates_file:
        with warnings.catch_warnings(record= True):
            warnings.simplefilter("always")
            Est = True
            Estimates_WB = openpyxl.load_workbook(Estimates_file, data_only=True)
            if JDE_check and hubble and Est:
                run_button.config(state= 'normal')
            else:
                run_button.config(state= 'disabled')
    else:
        label_file_explorer_estimates.configure(text="Error loading file, please try again", fg= "red")

def browseFiles_Actuals_Hubble():
    global Actuals_WB_Hubble
    global Est
    global hubble 
    global JDE_check
    hubble = False

    Actuals_file_Hubble = filedialog.askopenfilename(initialdir = "/", title = "Select a File",filetypes = (("Excel Workbooks", "*.xlsx*"),("all files", "*.*")))

    # Change label contents
    label_file_explorer_actuals_hubble.configure(text="File Opened: "+ Actuals_file_Hubble, fg = "blue")

    if Actuals_file_Hubble:
        with warnings.catch_warnings(record= True):
            warnings.simplefilter("always")
            hubble = True
            Actuals_WB_Hubble = openpyxl.load_workbook(Actuals_file_Hubble)
            if JDE_check and hubble and Est:
                run_button.config(state= 'normal')
            else:
                run_button.config(state= 'disabled')
    else:
        label_file_explorer_actuals_hubble.configure(text="Error loading file, please try again", fg= "red")

def browseFiles_Actuals_JDE():
    global Actuals_WB_JDE
    global Est
    global hubble 
    global JDE_check
    JDE_check = False

    Actuals_file = filedialog.askopenfilename(initialdir = "/", title = "Select a File",filetypes = (("Excel Workbooks", "*.xlsx*"),("all files", "*.*")))

    # Change label contents
    label_file_explorer_actuals_JDE.configure(text="File Opened: "+ Actuals_file, fg = "blue")

    if Actuals_file:
        with warnings.catch_warnings(record= True):
            warnings.simplefilter("always")
            JDE_check = True
            Actuals_WB_JDE = openpyxl.load_workbook(Actuals_file)
            if JDE_check == True and hubble and Est:
                run_button.config(state= 'normal')
            else:
                run_button.config(state= 'disabled')
    else:
        label_file_explorer_actuals_JDE.configure(text="Error loading file, please try again", fg= "red")


window = tkinter.Tk()
window.title("Project Variance Analysis")

frame = tkinter.Frame(window)
frame.pack() 

photo = PhotoImage(file="icon.png")
window.iconphoto(False, photo)

#reading estimate file
info_frame = tkinter.LabelFrame(frame, text="Work order Information")
info_frame.grid(row= 0, column= 0, padx = 20, pady= 10)

WO_number_label = tkinter.Label(info_frame, text= "Enter Work Order Number: ")
WO_number_label.grid(row= 0, column= 0)

WO_number_Entry = tkinter.Entry(info_frame)
WO_number_Entry.grid(row= 0, column= 1)

for widget in info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

file_frame = tkinter.LabelFrame(frame, text="Estimate data File upload")
file_frame.grid(row= 1, column= 0, sticky= "news", padx = 20, pady= 10)

Estimates_label = tkinter.Label(file_frame, text= "Upload Estimate file: ")
Estimates_label.grid(row= 0, column= 0)

Estimates_button = tkinter.Button(file_frame, text= "Upload File", command=browseFiles_Estimates)
Estimates_button.grid(row= 0, column= 1)

file_frame_actuals = tkinter.LabelFrame(frame, text="Actual data Files upload")
file_frame_actuals.grid(row= 2, column= 0, sticky= "news", padx = 20, pady= 10)

Actuals_label = tkinter.Label(file_frame_actuals, text= "Upload Hubble file: ")
Actuals_label.grid(row= 0, column= 0)
Actuals_label = tkinter.Label(file_frame_actuals, text= "Upload JDE WO costs file: ")
Actuals_label.grid(row= 1, column= 0)

Actuals_button = tkinter.Button(file_frame_actuals, text= "Upload File", command=browseFiles_Actuals_Hubble)
Actuals_button.grid(row= 0, column= 1)
Actuals_button = tkinter.Button(file_frame_actuals, text= "Upload File", command=browseFiles_Actuals_JDE)
Actuals_button.grid(row= 1, column= 1)

label_file_explorer_estimates = Label(file_frame, text = "",fg = "blue")
label_file_explorer_estimates.grid(row=0, column=2)

label_file_explorer_actuals_hubble = Label(file_frame_actuals, text = "",fg = "blue")
label_file_explorer_actuals_hubble.grid(row=0, column=2)

label_file_explorer_actuals_JDE = Label(file_frame_actuals, text = "",fg = "blue")
label_file_explorer_actuals_JDE.grid(row=1, column=2)

for widget in file_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

for widget in file_frame_actuals.winfo_children():
    widget.grid_configure(padx=10, pady=5)

run_button = tkinter.Button(frame, text="Run analysis", command = run_analysis, state= 'disabled')
run_button.grid(row=3, column=0, sticky="news", padx=20, pady=10)


window.mainloop()
